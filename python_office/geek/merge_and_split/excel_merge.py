import xlrd
import xlwt
from pathlib import Path, PurePath
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

'''
支持 Excel 读取的扩展库叫做 xlrd 库，支持 Excel 写入的扩展库叫做 xlwt 库
安装命令：
    pip3 install xlrd 
    pip3 install xlwt

所欲异常：
xlrd.biffh.XLRDError: Excel xlsx file; not supported
处理方式：
pip uninstall xlrd
pip install xlrd==1.2.0 (或者更早版本)

ps: 产生问题的根源是，xlrd这个软件包最新的版本不支持xlsx格式导致的，一般软件更新兼容性会更好，但是这个偏偏是个例，它的兼容性发生了倒退，所以我们使用更早期的版本来实现，xls为后缀的excel文件的读取
xlwt is a library for writing data and formatting information to older Excel files (ie: .xls)
如果要处理最新的xlsx后缀的excel,通常使用 openpyxl来进行操作


'''


def xl_read():
    path = "C:\\Users\\admin\\Desktop\\相关文档\\12月迭代功能完成情况统计表.xlsx"
    # path = "C:\\Users\\admin\\Desktop\\python.xlsx"
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    value = table.cell_value(rowx=1, colx=1)
    print(value)


def xl_write():
    # 使用相对路劲测试 xlwt保存的xlsx 后缀文件打不开，只能保存为xls后缀文件
    xl_path = "execl\\chenshaoxian.xls"
    # 创建实例
    workbook = xlwt.Workbook(encoding='utf-8')
    # 添加表sheet
    xlsheet = workbook.add_sheet('测试添加sheet')

    xlsheet.write(0, 0, "测试写入内容")
    workbook.save(xl_path)


def xls_files(src_dir):
    p = Path(src_dir)
    # 列表推导式的方式
    files = [x for x in p.iterdir() if PurePath(x).match('*.xls')]
    return files


def xlsx_files(src_dir):
    p = Path(src_dir)
    # 列表推导式的方式
    files = [x for x in p.iterdir() if PurePath(x).match('*.xlsx')]
    return files


content = []


def merge(src_path):
    dst_file = "execl/汇总文件.xls"
    files = xls_files(src_path)
    for file in files:
        data = xlrd.open_workbook(file)
        table = data.sheets()[0]
        username = table.cell_value(1, 0)
        score1 = table.cell_value(1, 2)
        score2 = table.cell_value(1, 2)
        # 字符串格式化拼接
        '''
        f-string，亦称为格式化字符串常量（formatted string literals），是Python3.6新引入的一种字符串格式化方法,主要目的是使格式化字符串的操作更加简便。
        f-string在形式上是以 f 或 F 修饰符引领的字符串,如（f'{xxx}' 或 F'{xxx}'），以大括号 {} 标明被替换的字段;f-string在本质上并不是字符串常量，而是一个在运行时运算求值的表达式
        '''
        temp = F'{username},{score1},{score1}'
        # 拆分拼装为二维数组
        content.append(temp.split(','))

    xls_header = ['姓名', '数学成绩', '语文成绩']
    workbook = xlwt.Workbook(encoding='utf-8')
    xls_sheet = workbook.add_sheet('成绩汇总')
    col = 0
    row = 0
    for cell_header in xls_header:
        xls_sheet.write(row, col, cell_header)
        col += 1

    row += 1
    for row_data in content:
        col = 0
        for cell in row_data:
            xls_sheet.write(row, col, cell)
            col += 1
        row += 1

    workbook.save(dst_file)


# 利用 openpyxl 库实现上面方法相同的功能逻辑
# openpyxl的行和列编号从1开始计算
def merge_xlsx(src_dir):
    dst_file = "execl/汇总文件.xlsx"
    files = xlsx_files(src_dir)
    # 读取各个excel中的原始数据，拆分拼装为二维数组
    for file in files:
        wb = load_workbook(file)
        # ws = wb.active # 获取当前活跃的worksheet,默认就是第一个worksheet 当然也可以使用下面的方法
        sheets = wb.sheetnames
        sheet_first = sheets[0]  # 第一个表格的名称
        # table = wb.get_sheet_by_name(sheet_first) 方法已经废弃
        table = wb[sheet_first]
        # 获取行和列
        rows = table.rows
        columns = table.columns
        list_rows = list(rows)
        size = len(list_rows)
        for index in range(size):
            if (index == 0):
                continue
            line = [col.value for col in list_rows[index]]
            content.append(line)

    print(content)
    # 写数据
    xlsx_header = ['姓名', '数学成绩', '语文成绩']
    workbook = Workbook()
    xlsx_sheet = workbook.active
    col = 1
    row = 1
    for cell_header in xlsx_header:
        xlsx_sheet.cell(row, col).value = cell_header
        col += 1

    row += 1
    for row_data in content:
        col = 1
        for cell in row_data:
            xlsx_sheet.cell(row, col).value = cell
            col += 1
        row += 1

    workbook.save(dst_file)


# merge('execl')
# merge_xlsx('execl')

def split(prefix, src_file):
    p = Path(prefix + '/' + src_file)
    wb = load_workbook(p)
    sheets = wb.sheetnames
    wb_sheet = wb.active  # 当前激活的工作簿 实测为当前正在编辑使用的
    row_datas = list(wb_sheet.values)
    print(row_datas)
    # sheet_first_name = sheets[0]
    # table = wb[sheet_first_name]  指定sheet名称获取 wb 对象
    rows = wb_sheet.rows
    columns = wb_sheet.columns
    max_row = wb_sheet.max_row  # 最大行数
    # file_name = wb_sheet.cell(1, 1).value 读取单元格数据

    # 表头处理
    xlsx_header = []
    for row in wb_sheet.iter_rows(max_row=1):
        xlsx_header = [cell.value for cell in row]
        print(xlsx_header)
    # 表数据处理
    for row in wb_sheet.iter_rows(min_row=2):
        content = [cell.value for cell in row]
        file_name = content[0]
        print(content)
        # 新文件写入
        workbook = Workbook()
        # 创建插入在 0 位置的sheet 这种方式总是会创建一个新的，导致表中有两个sheet
        # sheet = workbook.create_sheet(file_name, 0)
        sheet = workbook.active
        sheet.title = file_name
        for i in range(0, len(xlsx_header)):
            sheet.cell(1, i + 1, xlsx_header[i])

        for i in range(0, len(content)):
            sheet.cell(2, i + 1, content[i])
        workbook.save(prefix + '/' + file_name + '.xlsx')


split('execl', '汇总文件.xlsx')
